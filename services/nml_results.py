"""Query and export NML milk-quality results for Collections."""

from __future__ import annotations

import csv
import datetime as dt
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select

from services.cows_in_milk import cows_in_milk_for_dates
from services.database import NmlMilkResult, get_session
from services.farms import FARMS

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

_HEADERS = (
    "Farm",
    "Producer Ref",
    "Sample Date",
    "Load",
    "Sample ID",
    "NML",
    "Litres",
    "Weighbridge",
    "Temp °C",
    "Cows in milk",
    "Butterfat %",
    "Protein %",
    "SCC",
    "BactoScan",
    "FPD",
    "A/B",
    "Urea",
)

_TREND_METRICS = ("butterfat_pct", "protein_pct", "scc", "bactoscan", "urea_pct", "fpd", "temp_c")
_FARM_ORDER = tuple(farm.code for farm in FARMS)


def _normalise_farms(farms: list[str] | None) -> list[str]:
    if not farms:
        return list(_FARM_ORDER)
    selected = {part.strip().upper() for part in farms if part and part.strip()}
    return [code for code in _FARM_ORDER if code in selected] or list(_FARM_ORDER)


def _ab_label(value: bool | None) -> str:
    if value is True:
        return "Pass"
    if value is False:
        return "Fail"
    return ""


def _is_nml_matched(row: NmlMilkResult) -> bool:
    return bool(row.nml_matched)


def _nml_status(row: NmlMilkResult, *, today: dt.date | None = None) -> str:
    return "matched" if _is_nml_matched(row) else "unmatched"


def _row_to_dict(row: NmlMilkResult, *, today: dt.date | None = None) -> dict[str, Any]:
    sample_missing = bool(row.sample_missing)
    status = _nml_status(row)
    return {
        "id": row.id,
        "farm": row.farm or "",
        "producer_ref": row.producer_ref or "",
        "milk_buyer": row.milk_buyer or "",
        "sample_date": row.sample_date.isoformat() if row.sample_date else "",
        "load_number": row.load_number,
        "sample_id": "" if sample_missing else (row.sample_id or ""),
        "sample_missing": sample_missing,
        "litres_load": row.litres_load,
        "litres_weighbridge": row.litres_weighbridge,
        "cows_in_milk": None,
        "temp_c": None if row.temp_c is None else round(float(row.temp_c), 2),
        "butterfat_pct": row.butterfat_pct,
        "protein_pct": row.protein_pct,
        "scc": row.scc,
        "bactoscan": row.bactoscan,
        "fpd": row.fpd,
        "antibiotic_pass": row.antibiotic_pass,
        "urea_pct": row.urea_pct,
        "report_month": row.report_month or "",
        "source": row.source or "",
        "matched": status == "matched",
        "nml_status": status,
    }


def _load_weight(row: dict[str, Any] | Any) -> float | None:
    """Litres used to weight quality averages. Prefer tanker litres, then weighbridge."""
    getter = row.get if isinstance(row, dict) else lambda key, default=None: getattr(row, key, default)
    for key in ("litres_load", "litres_weighbridge"):
        raw = getter(key)
        if raw is None:
            continue
        try:
            weight = float(raw)
        except (TypeError, ValueError):
            continue
        if weight > 0:
            return weight
    return None


def _weighted_avg(rows: list[Any], metric: str, digits: int = 2) -> float | None:
    total = 0.0
    weight = 0.0
    for row in rows:
        getter = row.get if isinstance(row, dict) else lambda key, default=None: getattr(row, key, default)
        value = getter(metric)
        litres = _load_weight(row)
        if value is None or litres is None:
            continue
        total += float(value) * litres
        weight += litres
    if weight <= 0:
        return None
    return round(total / weight, digits)


def _attach_cows_in_milk(session: Any, rows: list[dict[str, Any]]) -> None:
    """Set cows_in_milk from herd inventory / events (same farm + date on every load)."""
    if not rows:
        return
    farms: set[str] = set()
    dates: set[dt.date] = set()
    parsed: list[tuple[dict[str, Any], str, dt.date | None]] = []
    for row in rows:
        farm = (row.get("farm") or "").strip().upper()
        day: dt.date | None = None
        raw_date = row.get("sample_date") or ""
        if isinstance(raw_date, dt.date):
            day = raw_date
        else:
            try:
                day = dt.date.fromisoformat(str(raw_date)[:10])
            except ValueError:
                day = None
        parsed.append((row, farm, day))
        if farm and day is not None:
            farms.add(farm)
            dates.add(day)
    counts = cows_in_milk_for_dates(session, farms, dates)
    for row, farm, day in parsed:
        if not farm or day is None:
            row["cows_in_milk"] = None
            continue
        row["cows_in_milk"] = counts.get((farm, day))


def _build_trend(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    buckets: dict[tuple[str, str], dict[str, list[tuple[float, float]]]] = {}
    day_litres: dict[tuple[str, str], float] = {}
    day_cows: dict[tuple[str, str], int] = {}
    for row in rows:
        farm = row["farm"] or "?"
        date = row["sample_date"]
        if not date:
            continue
        key = (farm, date)
        litres = _load_weight(row)
        if litres:
            day_litres[key] = day_litres.get(key, 0.0) + litres
        cows = row.get("cows_in_milk")
        if cows is not None:
            try:
                day_cows[key] = int(cows)
            except (TypeError, ValueError):
                pass
        if litres is None:
            continue
        bucket = buckets.setdefault(key, {m: [] for m in _TREND_METRICS})
        for metric in _TREND_METRICS:
            value = row.get(metric)
            if value is not None:
                bucket[metric].append((float(value), litres))

    trend: dict[str, list[dict[str, Any]]] = {}
    for (farm, date), metrics in buckets.items():
        litres = round(day_litres.get((farm, date), 0.0), 0)
        cows = day_cows.get((farm, date))
        point: dict[str, Any] = {
            "date": date,
            "litres": litres or None,
            "volume_litres": litres or None,
            "cows_in_milk": cows,
        }
        if litres and cows and cows > 0:
            point["litres_per_cow"] = round(float(litres) / cows, 2)
        else:
            point["litres_per_cow"] = None
        for metric in _TREND_METRICS:
            pairs = metrics[metric]
            if not pairs:
                point[metric] = None
                continue
            total = sum(value * weight for value, weight in pairs)
            weight = sum(weight for _value, weight in pairs)
            point[metric] = round(total / weight, 3) if weight else None
        trend.setdefault(farm, []).append(point)
    for points in trend.values():
        points.sort(key=lambda item: item["date"])
    return trend


def _summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    latest = max((r["sample_date"] for r in rows if r["sample_date"]), default="")
    fails = sum(1 for r in rows if r.get("antibiotic_pass") is False)
    litres = [float(r["litres_load"]) for r in rows if r.get("litres_load") is not None]
    weighbridge = [
        float(r["litres_weighbridge"])
        for r in rows
        if r.get("litres_weighbridge") is not None
    ]
    return {
        "count": len(rows),
        "latest_sample_date": latest,
        "avg_butterfat_pct": _weighted_avg(rows, "butterfat_pct"),
        "avg_protein_pct": _weighted_avg(rows, "protein_pct"),
        "avg_scc": _weighted_avg(rows, "scc"),
        "avg_bactoscan": _weighted_avg(rows, "bactoscan"),
        "avg_urea_pct": _weighted_avg(rows, "urea_pct", 3),
        "avg_temp_c": _weighted_avg(rows, "temp_c", 2),
        "antibiotic_fails": fails,
        "total_litres": round(sum(litres), 0) if litres else None,
        "total_weighbridge": round(sum(weighbridge), 0) if weighbridge else None,
        "missing_sample_count": sum(1 for r in rows if r.get("sample_missing")),
        "matched_count": sum(1 for r in rows if r.get("matched")),
        "nml_unmatched_count": sum(1 for r in rows if not r.get("matched")),
    }


def list_nml_results(
    *,
    farms: list[str] | None = None,
    date_from: dt.date | None = None,
    date_to: dt.date | None = None,
) -> dict[str, Any]:
    selected_farms = _normalise_farms(farms)
    with get_session() as session:
        query = select(NmlMilkResult).where(NmlMilkResult.farm.in_(selected_farms))
        if date_from is not None:
            query = query.where(NmlMilkResult.sample_date >= date_from)
        if date_to is not None:
            query = query.where(NmlMilkResult.sample_date <= date_to)
        query = query.order_by(
            NmlMilkResult.sample_date.desc(),
            NmlMilkResult.load_number.asc(),
            NmlMilkResult.sample_id.desc(),
        )
        rows = [_row_to_dict(row, today=dt.date.today()) for row in session.scalars(query).all()]
        _attach_cows_in_milk(session, rows)
    return {
        "rows": rows,
        "total": len(rows),
        "summary": _summary(rows),
        "trend": _build_trend(rows),
        "unmatched_nml": [],
    }


def nml_status() -> dict[str, Any]:
    with get_session() as session:
        row_count = session.scalar(select(func.count()).select_from(NmlMilkResult)) or 0
        latest_import = session.scalar(select(func.max(NmlMilkResult.imported_at)))
        earliest_sample = session.scalar(select(func.min(NmlMilkResult.sample_date)))
        latest_sample = session.scalar(select(func.max(NmlMilkResult.sample_date)))
        rows_by_farm = {
            farm: count
            for farm, count in session.execute(
                select(NmlMilkResult.farm, func.count()).group_by(NmlMilkResult.farm)
            ).all()
        }
    return {
        "row_count": row_count,
        "rows_by_farm": rows_by_farm,
        "latest_import": latest_import.isoformat() if latest_import else None,
        "earliest_sample_date": earliest_sample.isoformat() if earliest_sample else None,
        "latest_sample_date": latest_sample.isoformat() if latest_sample else None,
    }


def _nml_export_label(row: dict[str, Any]) -> str:
    status = row.get("nml_status") or ("matched" if row.get("matched") else "pending")
    if status == "matched":
        return "Matched"
    return "Unmatched"


def _export_cells(row: dict[str, Any]) -> list[Any]:
    return [
        row.get("farm", ""),
        row.get("producer_ref", ""),
        row.get("sample_date", ""),
        row.get("load_number"),
        row.get("sample_id", ""),
        _nml_export_label(row),
        row.get("litres_load"),
        row.get("litres_weighbridge"),
        row.get("temp_c"),
        row.get("cows_in_milk"),
        row.get("butterfat_pct"),
        row.get("protein_pct"),
        row.get("scc"),
        row.get("bactoscan"),
        row.get("fpd"),
        _ab_label(row.get("antibiotic_pass")),
        row.get("urea_pct"),
    ]


def build_nml_results_csv(rows: list[dict[str, Any]]) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(_HEADERS)
    for row in rows:
        writer.writerow(_export_cells(row))
    return buffer.getvalue()


def build_nml_results_xlsx(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "NML Results"
    ws.append(list(_HEADERS))
    for row in rows:
        ws.append(_export_cells(row))
    widths = [8, 14, 14, 8, 12, 10, 12, 12, 10, 12, 12, 11, 8, 11, 8, 8, 9]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
