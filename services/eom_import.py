"""Import historic milk collections from farm EOM workbooks (one row per load)."""

from __future__ import annotations

import calendar
import datetime as dt
import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select

from services.database import NmlMilkResult, get_session
from services.nml_pdf import PRODUCER_REF_FARM

logger = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent.parent
ALH_PRODUCER_REF = "641565"
DEFAULT_EOM_PATH = ROOT / "EOM Aston Lower Hall.xlsx"
ALLOWED_MONTHS = {(2026, 4), (2026, 5), (2026, 6), (2026, 7), (2026, 8)}

# Parkes Load 2 is milk fed to calves, not a tanker collection (always well under 1000 L).
PRK_CALF_MILK_MAX_LITRES = 1000.0

# Filename -> farm is the source of truth for this backfill.
EOM_WORKBOOKS: tuple[dict[str, str], ...] = (
    {"file": "EOM Aston Lower Hall.xlsx", "farm": "ALH", "producer_ref": "641565"},
    {"file": "EOM Bank.xlsx", "farm": "BNK", "producer_ref": "618538"},
    {"file": "EOM Cherry Orchard.xlsx", "farm": "COF", "producer_ref": "930221"},
    {"file": "EOM Park Hall.xlsx", "farm": "SFR", "producer_ref": "527634"},
    {"file": "EOM Parkes.xlsx", "farm": "PRK", "producer_ref": "231000002"},
)

_MONTH_ALIASES = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}

_METRICS = ("Milk", "Fat", "Protein", "SCC", "BCC", "Urea", "Sample No")
_COL_N = 14
_COL_O = 15
_COL_P = 16

_UPSERT_FIELDS = (
    "farm",
    "milk_buyer",
    "report_month",
    "report_date",
    "load_number",
    "litres_load",
    "litres_weighbridge",
    "butterfat_pct",
    "protein_pct",
    "scc",
    "bactoscan",
    "fpd",
    "antibiotic_pass",
    "urea_pct",
    "sample_missing",
    "source",
    "source_file",
)


def _parse_sheet_month(name: str) -> tuple[int, int] | None:
    parts = name.strip().split()
    if len(parts) != 2 or parts[0].lower().startswith("annual"):
        return None
    month = _MONTH_ALIASES.get(parts[0].lower())
    if month is None or not parts[1].isdigit():
        return None
    year_part = int(parts[1])
    year = year_part if year_part >= 100 else 2000 + year_part
    return year, month


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text or text.startswith("#"):
            return None
        try:
            return float(text.replace(",", ""))
        except ValueError:
            return None
    if _is_number(value):
        return float(value)
    return None


def _as_int(value: Any) -> int | None:
    number = _as_float(value)
    if number is None:
        return None
    return int(round(number))


def _as_percent(value: Any) -> float | None:
    number = _as_float(value)
    if number is None or number == 0:
        return None
    if abs(number) <= 1:
        return round(number * 100.0, 4)
    if abs(number) <= 15:
        return round(number, 4)
    return None


def _scale_urea(farm: str, value: Any) -> float | None:
    """SFR/COF EOM sheets store urea 1000x too high (22 instead of 0.022)."""
    number = _as_float(value)
    if number is None:
        return None
    if farm in {"SFR", "COF"} and abs(number) >= 1:
        return round(number / 1000.0, 6)
    return number


def _as_sample(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if _is_number(value) and float(value) == 0:
        return None
    if _is_number(value) and float(value) == int(value):
        return str(int(value))
    if _is_number(value):
        return str(value).rstrip("0").rstrip(".")
    text = str(value).strip()
    if not text or text.startswith("#"):
        return None
    return text


def _cell(ws, row: int, col: int):
    return ws.cell(row, col).value


def _report_month_label(year: int, month: int) -> str:
    return f"{calendar.month_name[month]} {year}"


def _is_prk_calf_milk(farm: str, litres: float | None) -> bool:
    return farm == "PRK" and litres is not None and 0 < litres < PRK_CALF_MILK_MAX_LITRES


def _detect_load_columns(ws) -> tuple[list[tuple[int, int]], dict[int, int]]:
    """First Collections 'Load 1..n' run at/after column P, plus Weight n if present."""
    header_row = 3
    for row in range(1, 6):
        for col in range(_COL_P, 40):
            if _header_text(_cell(ws, row, col)) == "load 1":
                header_row = row
                break
        else:
            continue
        break

    first_load = None
    for col in range(_COL_P, 40):
        if _header_text(_cell(ws, header_row, col)) == "load 1":
            first_load = col
            break
    if first_load is None:
        return [], {}

    loads: list[tuple[int, int]] = []
    load_number = 1
    col = first_load
    while col <= 40:
        if _header_text(_cell(ws, header_row, col)) == f"load {load_number}":
            loads.append((load_number, col))
            load_number += 1
            col += 1
            continue
        break

    weights: dict[int, int] = {}
    first_weight = None
    for col in range(first_load, 45):
        if _header_text(_cell(ws, header_row, col)) == "weight 1":
            first_weight = col
            break
    if first_weight is not None:
        weight_number = 1
        col = first_weight
        while col <= 45:
            if _header_text(_cell(ws, header_row, col)) == f"weight {weight_number}":
                weights[weight_number] = col
                weight_number += 1
                col += 1
                continue
            break
    return loads, weights


def parse_eom_loads(
    path: Path,
    *,
    producer_ref: str = ALH_PRODUCER_REF,
    farm: str | None = None,
    date_from: dt.date | None = None,
    allowed_months: set[tuple[int, int]] | None = None,
) -> list[dict[str, Any]]:
    farm_code = (farm or PRODUCER_REF_FARM.get(producer_ref) or "ALH").upper()
    cutoff = date_from or dt.date(2026, 4, 1)
    months = allowed_months if allowed_months is not None else ALLOWED_MONTHS
    wb = load_workbook(path, data_only=True, read_only=False)
    records: list[dict[str, Any]] = []
    used_ids: set[tuple[str, dt.date, str]] = set()

    try:
        for sheet_name in wb.sheetnames:
            parsed = _parse_sheet_month(sheet_name)
            if parsed is None:
                continue
            year, month = parsed
            if months is not None and (year, month) not in months:
                continue
            last_day = calendar.monthrange(year, month)[1]
            if dt.date(year, month, last_day) < cutoff:
                continue
            ws = wb[sheet_name]
            load_cols, weight_cols = _detect_load_columns(ws)
            if not load_cols:
                logger.warning("No Load columns on %s [%s]", path.name, sheet_name)
                continue
            report_month = _report_month_label(year, month)
            row = 1
            max_row = ws.max_row or 1
            while row <= max_row:
                metric = _cell(ws, row, _COL_O)
                if not (isinstance(metric, str) and metric.strip().lower() == "milk"):
                    row += 1
                    continue
                day_raw = _as_int(_cell(ws, row, _COL_N))
                if day_raw is None or day_raw < 1 or day_raw > last_day:
                    row += 7
                    continue
                sample_date = dt.date(year, month, day_raw)
                if sample_date < cutoff:
                    row += 7
                    continue

                metric_rows = {name: row + index for index, name in enumerate(_METRICS)}
                for load_number, load_col in load_cols:
                    litres = _as_float(_cell(ws, metric_rows["Milk"], load_col))
                    if litres is None or litres <= 0:
                        continue
                    if _is_prk_calf_milk(farm_code, litres):
                        continue
                    weight_col = weight_cols.get(load_number)
                    sample = _as_sample(_cell(ws, metric_rows["Sample No"], load_col))
                    if not sample and weight_col:
                        sample = _as_sample(_cell(ws, metric_rows["Sample No"], weight_col))
                    sample_missing = not bool(sample)
                    if sample_missing:
                        sample = f"L{load_number}"
                    key = (producer_ref, sample_date, sample)
                    if key in used_ids:
                        sample = f"{sample}-L{load_number}"
                        sample_missing = True
                    used_ids.add((producer_ref, sample_date, sample))

                    records.append(
                        {
                            "producer_ref": producer_ref,
                            "farm": farm_code,
                            "milk_buyer": None,
                            "report_month": report_month,
                            "report_date": None,
                            "sample_date": sample_date,
                            "sample_id": sample,
                            "load_number": load_number,
                            "litres_load": litres,
                            "litres_weighbridge": (
                                _as_float(_cell(ws, metric_rows["Milk"], weight_col))
                                if weight_col
                                else None
                            ),
                            "butterfat_pct": _as_percent(
                                _cell(ws, metric_rows["Fat"], load_col)
                            ),
                            "protein_pct": _as_percent(
                                _cell(ws, metric_rows["Protein"], load_col)
                            ),
                            "scc": _as_int(_cell(ws, metric_rows["SCC"], load_col)),
                            "bactoscan": _as_int(_cell(ws, metric_rows["BCC"], load_col)),
                            "fpd": None,
                            "antibiotic_pass": None,
                            "urea_pct": _scale_urea(
                                farm_code, _cell(ws, metric_rows["Urea"], load_col)
                            ),
                            "sample_missing": sample_missing,
                            "source": "eom",
                            "source_file": path.name,
                        }
                    )
                row += 7
    finally:
        wb.close()

    records.sort(key=lambda item: (item["sample_date"], item["load_number"]))
    return records


def delete_prk_calf_milk_rows() -> int:
    """Remove Parkes calf-milk loads (under 1000 L) from collections."""
    with get_session() as session:
        rows = session.scalars(select(NmlMilkResult).where(NmlMilkResult.farm == "PRK")).all()
        deleted = 0
        for row in rows:
            if _is_prk_calf_milk("PRK", row.litres_load):
                session.delete(row)
                deleted += 1
        session.commit()
    return deleted


def import_eom_collections(
    path: Path | None = None,
    *,
    producer_ref: str = ALH_PRODUCER_REF,
    farm: str | None = None,
    date_from: dt.date | None = None,
) -> dict[str, Any]:
    workbook = Path(path) if path else DEFAULT_EOM_PATH
    if not workbook.is_file():
        raise FileNotFoundError(f"EOM workbook not found: {workbook}")

    farm_code = (farm or PRODUCER_REF_FARM.get(producer_ref) or "ALH").upper()
    records = parse_eom_loads(
        workbook,
        producer_ref=producer_ref,
        farm=farm_code,
        date_from=date_from,
    )
    inserted, updated = _upsert(records)
    calf_removed = 0
    if farm_code == "PRK":
        calf_removed = delete_prk_calf_milk_rows()
    missing = sum(1 for row in records if row["sample_missing"])
    return {
        "source_file": workbook.name,
        "rows_parsed": len(records),
        "rows_inserted": inserted,
        "rows_updated": updated,
        "rows_missing_sample": missing,
        "rows_calf_milk_removed": calf_removed,
        "date_from": (date_from or dt.date(2026, 4, 1)).isoformat(),
        "date_to": records[-1]["sample_date"].isoformat() if records else None,
        "farm": farm_code,
    }


def import_all_eom_workbooks(
    *,
    directory: Path | None = None,
    date_from: dt.date | None = None,
) -> list[dict[str, Any]]:
    folder = directory or ROOT
    results: list[dict[str, Any]] = []
    for item in EOM_WORKBOOKS:
        path = folder / item["file"]
        if not path.is_file():
            logger.warning("EOM workbook missing: %s", path)
            results.append(
                {
                    "source_file": item["file"],
                    "farm": item["farm"],
                    "error": "file not found",
                }
            )
            continue
        results.append(
            import_eom_collections(
                path,
                producer_ref=item["producer_ref"],
                farm=item["farm"],
                date_from=date_from,
            )
        )
    return results


def _upsert(records: list[dict[str, Any]]) -> tuple[int, int]:
    if not records:
        return (0, 0)

    producer_refs = {row["producer_ref"] for row in records}
    with get_session() as session:
        existing_rows = session.scalars(
            select(NmlMilkResult).where(NmlMilkResult.producer_ref.in_(producer_refs))
        ).all()
        by_sample = {
            (row.producer_ref, row.sample_date, row.sample_id): row for row in existing_rows
        }
        by_load = {
            (row.producer_ref, row.sample_date, row.load_number): row
            for row in existing_rows
            if row.load_number is not None
        }

        inserted = 0
        updated = 0
        for record in records:
            load_key = (record["producer_ref"], record["sample_date"], record["load_number"])
            sample_key = (record["producer_ref"], record["sample_date"], record["sample_id"])
            row = by_load.get(load_key) or by_sample.get(sample_key)
            if row is None:
                row = NmlMilkResult(**record)
                session.add(row)
                by_load[load_key] = row
                by_sample[sample_key] = row
                inserted += 1
                continue
            for field in _UPSERT_FIELDS:
                setattr(row, field, record.get(field))
            row.sample_id = record["sample_id"]
            row.imported_at = dt.datetime.now(dt.timezone.utc)
            updated += 1
        session.commit()
    return (inserted, updated)
